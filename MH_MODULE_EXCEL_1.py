import os,sys
sys.path.append(os.path.dirname(__file__))
from MH_MODULE_SQLQUERIES_1 import *

# XLS Management libraries  
from openpyxl import Workbook 
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont 

couleur_homme ="#ffcc99"
couleur_femme="#ccccff"
Black ="#000000"
font_texte ="arial narrow"

#### EXCEL SECTION
def init_sheet():
#-------------------------------------------------------------------------------------------------------------
    wb = Workbook() 
    ws_arbre = wb['Sheet'] 
    ws_arbre.title = 'Arbre'
#-------------------------------------------------------------------------------------------------------------
    return wb,ws_arbre
#=============================================================================================================
def excel_arbre(sql_obj,ws_arbre,la_cible_id,n_levels):

    MH_cible = get_personne_by_indi_id(sql_obj,la_cible_id)
    MH_ascendants = get_personne_ascendants(sql_obj,MH_cible,[],0,n_levels)
#-------------------------------------------------------------------------------------------------------------
    #MH_ascendants = [n_level= 0 ,MH_adult1= 1 ,MH_adult1 = 2 ,False = 3 ,n_rank = 4, side = 5 ]
#-------------------------------------------------------------------------------------------------------------    
    ws_arbre.sheet_view.showGridLines = False
    i_arbre_row = 2
    for MH_ascendant in MH_ascendants:
        # box position  
        i_arbre_col = MH_ascendant.level * 3

        MH_personne = MH_ascendant.adult1
        if MH_personne.sexe == "F" : i_arbre_row = i_arbre_row + 1

        #add cell to  Arbre   
        excel_arbre_dessine_box(sql_obj,ws_arbre,i_arbre_row,i_arbre_col,MH_personne)

    #Resize Arbre Colonne
    excel_arbre_polish(ws_arbre,12,30)
    return
#=============================================================================================================
def excel_arbre_dessine_box(sql_obj,sheet,row,col,MH_personne):
#-----------------------------------------------------------------------------------------------------------------
    le_text = CellRichText(
        TextBlock(font=InlineFont(sz=14,rFont=font_texte,b=True,color=Black[1:]),
                                    text=f'{MH_personne.prenom}\n'),                                
        TextBlock(font=InlineFont(sz=14,rFont=font_texte,b=True,color=Black[1:]),
                                    text=f'{MH_personne.nom}\n'), 
        TextBlock(font=InlineFont(sz=12,rFont=font_texte,b=False,color=Black[1:]),
                                    text=f'{MH_personne.prenoms if MH_personne.prenoms else "" }{", "+ MH_personne.surnom if MH_personne.surnom else ""}\n'), 
        TextBlock(font=InlineFont(sz=12,rFont=font_texte,b=False,i=False,color=Black[1:]),
                                    text=f'{MH_personne.bdate}')                                                                       
        )

    sheet.cell(row=row, column=col).value = le_text      

    #cell alignement
    sheet.cell(row=row, column=col).alignment =Alignment(horizontal='center',
                                                vertical='center',
                                                text_rotation=0,
                                                wrap_text=True,
                                                shrink_to_fit=True,
                                                indent=0)
    #cell fill
    if MH_personne.sexe == "M":
        fill_color = couleur_homme[1:]
    else:
        fill_color = couleur_femme[1:]

    sheet.cell(row=row, column=col).fill =  PatternFill(fill_type='solid',
                                                    start_color=fill_color,
                                                    end_color=fill_color)
    if col > 1:
        sheet.cell(row=row, column=col-1).fill =  PatternFill(fill_type='solid',
                                                        start_color=fill_color,
                                                        end_color=fill_color)
        sheet.cell(row=row, column=col-1).fill =  PatternFill(fill_type='solid',
                                                        start_color=fill_color,
                                                        end_color=fill_color)

#--- Add photos-------------------------------------------------------------------------------------------------------------------
    try:
        img_path = get_personne_photoID_file(sql_obj,MH_personne)
        excel_add_image_sheet(img_path,sheet,row,col)
    except Exception as error:
        print(error)
             
    return
#=============================================================================================================
def excel_add_image_sheet(img_path,sheet,row,col):
    from openpyxl.drawing.image import Image
    from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker 
    from openpyxl.drawing.xdr import XDRPositiveSize2D
        #insert image
    img = Image(img_path)
    r = img.height /img.width
    h = 90
    w = h / r
    if w > 90: w = 90
    img.height = h
    img.width= w

    c2e = cm_to_EMU
    # Calculated number of cells width or height from cm into EMUs
    cellh = lambda x: c2e((x * 49.77)/99)
    cellw = lambda x: c2e((x * (18.65-1.71))/10)

    # Also offset 
    coloffset = cellw(0.2)
    rowoffset =  cellh(0)

    p2e = pixels_to_EMU
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-2, colOff=coloffset, row=row-1, rowOff=rowoffset)
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    sheet.add_image(img)
    return
#=============================================================================================================
def excel_arbre_polish(sheet,size_image,size_texte):
    le_max_col = sheet.max_column
    le_max_row = sheet.max_row
    for i in range(2,le_max_col,3):    
        sheet.column_dimensions[excel_column_name(i)].width=size_image  
        i += 1
        sheet.column_dimensions[excel_column_name(i)].width=size_texte     
        i += 1
        sheet.column_dimensions[excel_column_name(i)].width=size_image 

        sheet.cell(row=1, column=i+1).value =  f"Ascendant niveau {int((i+1)/3)}"       

#   cell border Cible or Frere/Soeur
    for i in range(0,le_max_col):
        sheet.cell(row=1, column=i+1).border =  Border(bottom=Side(border_style="thick",color=Black[1:]))

# marge colorée
    for i in range(1,le_max_row+1):
        sheet.cell(row=i, column=1).fill =  PatternFill(fill_type='solid',start_color="EFEFEF",end_color="EFEFEF") 
        if i > 1: sheet.row_dimensions[i].height = 75
    return
#=============================================================================================================
def excel_column_name(n):
    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    name = ''
    while n > 0:
        n, r = divmod (n - 1, 26)
        name = chr(r + ord('A')) + name
    return name
#=============================================================================================================
def excel_column_number(name):
#-------------------------------------------------------------------------------------------------------------  
    """Excel-style column name to number, e.g., A = 1, Z = 26, AA = 27, AAA = 703."""
    n = 0
    for c in name:
        n = n * 26 + 1 + ord(c) - ord('A')
#-------------------------------------------------------------------------------------------------------------  
    return n        
#=============================================================================================================
def excel_ascendants(MH_ascendants,n_level,n_rank,MH_cible,le_side):
#-------------------------------------------------------------------------------------------------------------
    if MH_cible:
        n_level = n_level + 1
        #if MH_cible.sex  == "F": n_rank = n_rank + 2
    #---------------------------------------------------------------------------------------------------------
        if n_level == 1: MH_ascendants.append([n_level,MH_cible,"???",False,n_rank,le_side]) 
    #---------------------------------------------------------------------------------------------------------
        MH_father = MH_cible.father
        MH_mother = MH_cible.mother

        if not MH_father and not MH_mother and MH_cible.sex == "F": n_rank = n_rank + 1

        if MH_father : 
            if le_side == "Cible" : le_side_P = "Paternel"
            else:le_side_P = le_side
            if MH_mother : MH_ascendants.append([n_level+1,MH_father,MH_mother,False,n_rank,le_side])
            else : MH_ascendants.append([n_level+1,MH_father,"???",False,n_rank+1,le_side])
            MH_ascendants,n_rank = excel_ascendants(MH_ascendants,n_level,n_rank,MH_father,le_side_P)

        if MH_mother : 
            n_rank = n_rank + 2
            if le_side == "Cible": le_side_M = "Maternel"
            else:le_side_M = le_side
            if MH_father : MH_ascendants.append([n_level+1,MH_mother,MH_father,False,n_rank,le_side])
            else : MH_ascendants.append([n_level+1,MH_mother,"???",False,n_rank+1,le_side])
            MH_ascendants,n_rank = excel_ascendants(MH_ascendants,n_level,n_rank,MH_mother,le_side_M)   
            
    return MH_ascendants,n_rank
#=============================================================================================================
def excel_column_name(n):
    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    name = ''
    while n > 0:
        n, r = divmod (n - 1, 26)
        name = chr(r + ord('A')) + name
    return name
#=============================================================================================================
def excel_column_number(name):
#-------------------------------------------------------------------------------------------------------------  
    """Excel-style column name to number, e.g., A = 1, Z = 26, AA = 27, AAA = 703."""
    n = 0
    for c in name:
        n = n * 26 + 1 + ord(c) - ord('A')
#-------------------------------------------------------------------------------------------------------------  
    return n        
#=============================================================================================================